"""Tests for Epic 3: Compliance Reporting — SOC 2, ISO 27001, NIST CSF, XLSX, scheduling."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_row(**kwargs: object) -> MagicMock:
    """Create a MagicMock that behaves like a SQLAlchemy Row with named attrs."""
    row = MagicMock()
    for key, value in kwargs.items():
        setattr(row, key, value)
    return row


def _mock_session(results: list | None = None) -> AsyncMock:
    """Return an ``AsyncSession`` mock with a default empty result set."""
    session = AsyncMock()
    mock_result = MagicMock()
    mock_result.fetchall.return_value = results or []
    mock_result.fetchone.return_value = (results or [None])[0] if results else _make_row(cnt=0)
    session.execute = AsyncMock(return_value=mock_result)
    return session


def _mock_session_multi(*result_sets: list) -> AsyncMock:
    """Return session mock that yields different result sets per execute() call."""
    session = AsyncMock()
    mocks = []
    for rows in result_sets:
        mock_result = MagicMock()
        mock_result.fetchall.return_value = rows
        mock_result.fetchone.return_value = rows[0] if rows else _make_row(cnt=0)
        mocks.append(mock_result)
    session.execute = AsyncMock(side_effect=mocks)
    return session


# ---------------------------------------------------------------------------
# SOC 2 Report Generation
# ---------------------------------------------------------------------------


class TestSOC2Report:
    """Tests for SOC 2 Type II evidence report generation."""

    @pytest.mark.asyncio
    async def test_generate_soc2_report_structure(self) -> None:
        """SOC 2 report should have correct top-level structure."""
        from app.services.compliance_report_service import generate_soc2_report

        # Each helper call returns a count row; we need many execute calls
        cnt_row = _make_row(cnt=5)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_soc2_report(session, now - timedelta(days=90), now, org=None)

        assert result["framework"] == "SOC 2 Type II"
        assert "period" in result
        assert "executive_summary" in result
        assert "controls" in result
        assert len(result["controls"]) == 5

    @pytest.mark.asyncio
    async def test_soc2_control_ids(self) -> None:
        """SOC 2 report should contain the correct control IDs."""
        from app.services.compliance_report_service import generate_soc2_report

        cnt_row = _make_row(cnt=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_soc2_report(session, now - timedelta(days=30), now)

        control_ids = [c["control_id"] for c in result["controls"]]
        assert control_ids == ["CC6.1", "CC6.2", "CC6.3", "CC8.1", "CC7.1"]

    @pytest.mark.asyncio
    async def test_soc2_with_org_filter(self) -> None:
        """SOC 2 report should pass org filter to queries."""
        from app.services.compliance_report_service import generate_soc2_report

        cnt_row = _make_row(cnt=3)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_soc2_report(session, now - timedelta(days=30), now, org="my-org")

        assert result["org"] == "my-org"
        # Verify org was passed — calls should include org param
        for call in session.execute.call_args_list:
            params = call[0][1] if len(call[0]) > 1 else {}
            if "org" in params:
                assert params["org"] == "my-org"

    @pytest.mark.asyncio
    async def test_soc2_executive_summary_fields(self) -> None:
        """Executive summary should contain expected metric keys."""
        from app.services.compliance_report_service import generate_soc2_report

        cnt_row = _make_row(cnt=10)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_soc2_report(session, now - timedelta(days=30), now)

        summary = result["executive_summary"]
        assert "total_audit_events" in summary
        assert "total_evidence_events" in summary
        assert "unique_actors" in summary
        assert summary["controls_assessed"] == 5

    @pytest.mark.asyncio
    async def test_soc2_period_calculation(self) -> None:
        """Period days should be calculated correctly."""
        from app.services.compliance_report_service import generate_soc2_report

        cnt_row = _make_row(cnt=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        start = datetime(2024, 1, 1, tzinfo=UTC)
        end = datetime(2024, 4, 1, tzinfo=UTC)
        result = await generate_soc2_report(session, start, end)

        assert result["period"]["days"] == 91


# ---------------------------------------------------------------------------
# ISO 27001 Report Generation
# ---------------------------------------------------------------------------


class TestISO27001Report:
    """Tests for ISO 27001 Annex A compliance report generation."""

    @pytest.mark.asyncio
    async def test_generate_iso27001_report_structure(self) -> None:
        """ISO 27001 report should have correct structure."""
        from app.services.compliance_report_service import generate_iso27001_report

        cnt_row = _make_row(cnt=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_iso27001_report(session, now - timedelta(days=90), now)

        assert result["framework"] == "ISO 27001 Annex A"
        assert len(result["controls"]) == 5

    @pytest.mark.asyncio
    async def test_iso27001_control_ids(self) -> None:
        """ISO 27001 should have the correct Annex A control IDs."""
        from app.services.compliance_report_service import generate_iso27001_report

        cnt_row = _make_row(cnt=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_iso27001_report(session, now - timedelta(days=30), now)

        control_ids = [c["control_id"] for c in result["controls"]]
        assert control_ids == ["A.9", "A.12", "A.14", "A.16", "A.18"]

    @pytest.mark.asyncio
    async def test_iso27001_compliance_score(self) -> None:
        """Compliance score should be between 0 and 100."""
        from app.services.compliance_report_service import generate_iso27001_report

        cnt_row = _make_row(cnt=5)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_iso27001_report(session, now - timedelta(days=30), now)

        a18 = result["controls"][4]  # A.18 is last
        assert a18["control_id"] == "A.18"
        score = a18["evidence"]["compliance_score_pct"]
        assert 0.0 <= score <= 100.0

    @pytest.mark.asyncio
    async def test_iso27001_executive_summary(self) -> None:
        """Executive summary should include compliance_score_pct."""
        from app.services.compliance_report_service import generate_iso27001_report

        cnt_row = _make_row(cnt=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_iso27001_report(session, now - timedelta(days=30), now)

        assert "compliance_score_pct" in result["executive_summary"]


# ---------------------------------------------------------------------------
# NIST CSF Report Generation
# ---------------------------------------------------------------------------


class TestNISTCSFReport:
    """Tests for NIST Cybersecurity Framework report generation."""

    @pytest.mark.asyncio
    async def test_generate_nist_csf_report_structure(self) -> None:
        """NIST CSF report should use 'functions' instead of 'controls'."""
        from app.services.compliance_report_service import generate_nist_csf_report

        cnt_row = _make_row(cnt=0, active_days=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_nist_csf_report(session, now - timedelta(days=90), now)

        assert result["framework"] == "NIST Cybersecurity Framework"
        assert "functions" in result
        assert len(result["functions"]) == 5

    @pytest.mark.asyncio
    async def test_nist_csf_function_ids(self) -> None:
        """NIST CSF should have ID, PR, DE, RS, RC functions."""
        from app.services.compliance_report_service import generate_nist_csf_report

        cnt_row = _make_row(cnt=0, active_days=0)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_nist_csf_report(session, now - timedelta(days=30), now)

        function_ids = [f["function_id"] for f in result["functions"]]
        assert function_ids == ["ID", "PR", "DE", "RS", "RC"]

    @pytest.mark.asyncio
    async def test_nist_csf_identify_evidence(self) -> None:
        """Identify function should contain asset inventory metrics."""
        from app.services.compliance_report_service import generate_nist_csf_report

        cnt_row = _make_row(cnt=42, active_days=25)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_nist_csf_report(session, now - timedelta(days=30), now)

        identify = result["functions"][0]
        assert identify["function_id"] == "ID"
        evidence = identify["evidence"]
        assert "unique_repositories" in evidence
        assert "unique_actors" in evidence
        assert "unique_organizations" in evidence

    @pytest.mark.asyncio
    async def test_nist_csf_recover_completeness(self) -> None:
        """Recover function should calculate audit trail completeness."""
        from app.services.compliance_report_service import generate_nist_csf_report

        cnt_row = _make_row(cnt=100, active_days=28)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(
                fetchall=MagicMock(return_value=[]),
                fetchone=MagicMock(return_value=cnt_row),
            )
        )

        now = datetime.now(UTC)
        result = await generate_nist_csf_report(session, now - timedelta(days=30), now)

        recover = result["functions"][4]
        assert recover["function_id"] == "RC"
        assert "audit_trail_completeness_pct" in recover["evidence"]


# ---------------------------------------------------------------------------
# XLSX Export
# ---------------------------------------------------------------------------


class TestXLSXService:
    """Tests for XLSX export generation."""

    def test_generate_xlsx_returns_bytes(self) -> None:
        """generate_xlsx should return bytes."""
        from app.services.xlsx_service import generate_xlsx

        data = [
            {"name": "Alice", "count": 10, "rate": 0.5},
            {"name": "Bob", "count": 20, "rate": 0.8},
        ]
        result = generate_xlsx(data, sheet_name="Test")

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_xlsx_valid_workbook(self) -> None:
        """Generated bytes should be a valid XLSX file."""
        import io

        from openpyxl import load_workbook

        from app.services.xlsx_service import generate_xlsx

        data = [
            {"bucket": "2024-01-01", "value": 42},
            {"bucket": "2024-01-02", "value": 99},
        ]
        result = generate_xlsx(data, sheet_name="Metrics")

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws is not None
        assert ws.title == "Metrics"
        # Header row + 2 data rows
        assert ws.max_row == 3

    def test_generate_xlsx_empty_data(self) -> None:
        """Empty data should produce a valid workbook with 'No data' message."""
        from app.services.xlsx_service import generate_xlsx

        result = generate_xlsx([], sheet_name="Empty")

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_xlsx_with_column_config(self) -> None:
        """Column config should control headers and order."""
        import io

        from openpyxl import load_workbook

        from app.services.xlsx_service import generate_xlsx

        data = [{"a": 1, "b": 2, "c": 3}]
        config = [
            {"key": "c", "header": "Column C"},
            {"key": "a", "header": "Column A"},
        ]
        result = generate_xlsx(data, sheet_name="Custom", column_config=config)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=1, column=1).value == "Column C"
        assert ws.cell(row=1, column=2).value == "Column A"

    def test_generate_xlsx_handles_none_values(self) -> None:
        """None values should be converted to empty strings."""
        from app.services.xlsx_service import generate_xlsx

        data = [{"name": None, "count": None}]
        result = generate_xlsx(data, sheet_name="Nulls")

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_xlsx_handles_nested_dicts(self) -> None:
        """Nested dicts should be serialized as JSON strings."""
        import io

        from openpyxl import load_workbook

        from app.services.xlsx_service import generate_xlsx

        data = [{"name": "test", "metadata": {"key": "value"}}]
        result = generate_xlsx(data, sheet_name="Nested")

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws is not None
        # Nested dict should be serialized
        cell_value = ws.cell(row=2, column=2).value
        assert "key" in str(cell_value)


# ---------------------------------------------------------------------------
# PDF / HTML Service
# ---------------------------------------------------------------------------


class TestPDFService:
    """Tests for HTML report rendering."""

    def test_render_compliance_report_html(self) -> None:
        """Should render a full HTML document from report data."""
        from app.services.pdf_service import render_compliance_report_html

        report = {
            "framework": "SOC 2 Type II",
            "generated_at": "2024-01-15T12:00:00+00:00",
            "period": {"start": "2024-01-01", "end": "2024-03-31", "days": 90},
            "org": None,
            "executive_summary": {
                "total_audit_events": 1000,
                "controls_assessed": 5,
            },
            "controls": [
                {
                    "control_id": "CC6.1",
                    "title": "Logical Access",
                    "description": "Test",
                    "evidence": {"role_changes": 50},
                    "status": "evidence_collected",
                }
            ],
            "functions": None,
        }

        html = render_compliance_report_html(report)

        assert "<!DOCTYPE html>" in html
        assert "SOC 2 Type II" in html
        assert "CC6.1" in html
        assert "Logical Access" in html

    def test_render_with_print_ready(self) -> None:
        """Print-ready mode should include @media print CSS."""
        from app.services.pdf_service import render_compliance_report_html

        report = {
            "framework": "Test",
            "generated_at": "2024-01-01",
            "period": {"start": "2024-01-01", "end": "2024-01-31", "days": 30},
            "org": None,
            "executive_summary": {"total_audit_events": 0},
            "controls": [],
            "functions": None,
        }

        html = render_compliance_report_html(report, print_ready=True)

        assert "@media print" in html
        assert "@page" in html

    def test_render_nist_csf_uses_functions(self) -> None:
        """NIST CSF reports should render function sections."""
        from app.services.pdf_service import render_compliance_report_html

        report = {
            "framework": "NIST CSF",
            "generated_at": "2024-01-01",
            "period": {"start": "2024-01-01", "end": "2024-01-31", "days": 30},
            "org": None,
            "executive_summary": {"total_audit_events": 0},
            "controls": None,
            "functions": [
                {
                    "function_id": "ID",
                    "title": "Identify",
                    "description": "Test",
                    "evidence": {"repos": 10},
                    "status": "evidence_collected",
                }
            ],
        }

        html = render_compliance_report_html(report)

        assert "Identify" in html
        assert "ID" in html


# ---------------------------------------------------------------------------
# Report Scheduling CRUD
# ---------------------------------------------------------------------------


class TestReportScheduleSchema:
    """Tests for report schedule Pydantic schemas."""

    def test_create_schema_valid(self) -> None:
        from app.schemas.report import ReportScheduleCreate

        s = ReportScheduleCreate(
            report_type="soc2",
            cron_expression="0 8 1 * *",
            export_format="html",
            recipients=["admin@example.com"],
        )
        assert s.report_type == "soc2"
        assert s.enabled is True

    def test_create_schema_invalid_format(self) -> None:
        from pydantic import ValidationError

        from app.schemas.report import ReportScheduleCreate

        with pytest.raises(ValidationError):
            ReportScheduleCreate(
                report_type="soc2",
                cron_expression="0 8 1 * *",
                export_format="docx",  # invalid
                recipients=[],
            )

    def test_update_schema_partial(self) -> None:
        from app.schemas.report import ReportScheduleUpdate

        s = ReportScheduleUpdate(enabled=False)
        dumped = s.model_dump(exclude_unset=True)
        assert dumped == {"enabled": False}

    def test_response_schema_from_attributes(self) -> None:
        """Response schema should support from_attributes for ORM objects."""
        from app.schemas.report import ReportScheduleResponse

        assert ReportScheduleResponse.model_config.get("from_attributes") is True


# ---------------------------------------------------------------------------
# Report Worker — cron parsing
# ---------------------------------------------------------------------------


class TestCronParsing:
    """Tests for the simplified cron matcher used by the report worker."""

    def test_cron_wildcard_matches(self) -> None:
        from app.workers.report_worker import _matches_field

        assert _matches_field("*", 0) is True
        assert _matches_field("*", 59) is True

    def test_cron_exact_match(self) -> None:
        from app.workers.report_worker import _matches_field

        assert _matches_field("30", 30) is True
        assert _matches_field("30", 31) is False

    def test_cron_step_value(self) -> None:
        from app.workers.report_worker import _matches_field

        assert _matches_field("*/15", 0) is True
        assert _matches_field("*/15", 15) is True
        assert _matches_field("*/15", 30) is True
        assert _matches_field("*/15", 7) is False

    def test_cron_comma_list(self) -> None:
        from app.workers.report_worker import _matches_field

        assert _matches_field("1,15", 1) is True
        assert _matches_field("1,15", 15) is True
        assert _matches_field("1,15", 2) is False

    def test_cron_range(self) -> None:
        from app.workers.report_worker import _matches_field

        assert _matches_field("1-5", 3) is True
        assert _matches_field("1-5", 0) is False
        assert _matches_field("1-5", 6) is False

    def test_cron_is_due_never_run(self) -> None:
        """A schedule that has never run should be due."""
        from app.workers.report_worker import _cron_is_due

        assert _cron_is_due("* * * * *", None) is True

    def test_cron_is_due_recently_run(self) -> None:
        """A schedule that ran recently should not be due again."""
        from app.workers.report_worker import _cron_is_due

        now = datetime.now(UTC)
        assert _cron_is_due("* * * * *", now - timedelta(minutes=10)) is False

    def test_cron_is_due_invalid_expression(self) -> None:
        """Invalid cron expression should return False."""
        from app.workers.report_worker import _cron_is_due

        assert _cron_is_due("bad cron", None) is False


# ---------------------------------------------------------------------------
# Report Worker — scheduled report task
# ---------------------------------------------------------------------------


class TestReportWorkerTask:
    """Tests for the scheduled report Celery task."""

    @pytest.mark.asyncio
    async def test_run_scheduled_reports_no_schedules(self) -> None:
        """With no schedules, should return 0 checked."""
        from app.workers.report_worker import _run_scheduled_reports

        mock_session = AsyncMock()
        mock_result = MagicMock()
        mock_result.scalars.return_value.all.return_value = []
        mock_session.execute = AsyncMock(return_value=mock_result)
        mock_session.commit = AsyncMock()

        with patch("app.workers.report_worker.AsyncSessionLocal") as mock_factory:
            mock_factory.return_value.__aenter__ = AsyncMock(return_value=mock_session)
            mock_factory.return_value.__aexit__ = AsyncMock(return_value=False)

            result = await _run_scheduled_reports()

        assert result["schedules_checked"] == 0
        assert result["reports_sent"] == 0


# ---------------------------------------------------------------------------
# Compliance report flattener
# ---------------------------------------------------------------------------


class TestFlattenComplianceReport:
    """Tests for flattening compliance reports to tabular format."""

    def test_flatten_soc2_controls(self) -> None:
        from app.workers.report_worker import _flatten_compliance_report

        report = {
            "controls": [
                {
                    "control_id": "CC6.1",
                    "title": "Access",
                    "evidence": {"events": 10, "actors": 5},
                }
            ]
        }
        rows = _flatten_compliance_report(report)
        assert len(rows) == 2
        assert rows[0]["section_id"] == "CC6.1"
        assert rows[0]["metric"] == "events"
        assert rows[0]["value"] == 10

    def test_flatten_nist_functions(self) -> None:
        from app.workers.report_worker import _flatten_compliance_report

        report = {
            "functions": [
                {
                    "function_id": "ID",
                    "title": "Identify",
                    "evidence": {"repos": 42},
                }
            ]
        }
        rows = _flatten_compliance_report(report)
        assert len(rows) == 1
        assert rows[0]["section_id"] == "ID"

    def test_flatten_empty_report(self) -> None:
        from app.workers.report_worker import _flatten_compliance_report

        rows = _flatten_compliance_report({})
        assert rows == []

    def test_flatten_non_numeric_values_as_strings(self) -> None:
        from app.workers.report_worker import _flatten_compliance_report

        report = {
            "controls": [
                {
                    "control_id": "A.9",
                    "title": "Access",
                    "evidence": {"top_actions": [{"action": "test", "count": 5}]},
                }
            ]
        }
        rows = _flatten_compliance_report(report)
        assert len(rows) == 1
        # Lists should be stringified
        assert isinstance(rows[0]["value"], str)


# ---------------------------------------------------------------------------
# Router integration tests
# ---------------------------------------------------------------------------


class TestComplianceRouterIntegration:
    """Tests for compliance report API endpoints via TestClient."""

    @pytest.mark.asyncio
    async def test_parse_date_range_defaults(self) -> None:
        """Default date range should use 90-day window."""
        from app.routers.reports import _parse_date_range

        start, end = _parse_date_range(None, None)
        diff = (end - start).days
        assert diff == 90

    @pytest.mark.asyncio
    async def test_parse_date_range_custom(self) -> None:
        """Custom ISO dates should be parsed correctly."""
        from app.routers.reports import _parse_date_range

        start, end = _parse_date_range("2024-01-01", "2024-03-31")
        assert start.year == 2024
        assert start.month == 1
        assert end.month == 3

    @pytest.mark.asyncio
    async def test_parse_date_range_with_timezone(self) -> None:
        """Dates with timezone info should be handled."""
        from app.routers.reports import _parse_date_range

        start, end = _parse_date_range(
            "2024-01-01T00:00:00+00:00",
            "2024-03-31T23:59:59+00:00",
        )
        assert start.tzinfo is not None
        assert end.tzinfo is not None

    def test_flatten_compliance_report_function(self) -> None:
        """Router's _flatten_compliance_report should work correctly."""
        from app.routers.reports import _flatten_compliance_report

        report = {
            "controls": [
                {
                    "control_id": "CC6.1",
                    "title": "Test",
                    "evidence": {"metric_a": 100},
                }
            ]
        }
        rows = _flatten_compliance_report(report)
        assert len(rows) == 1
        assert rows[0]["section_id"] == "CC6.1"


# ---------------------------------------------------------------------------
# Report catalog includes compliance reports
# ---------------------------------------------------------------------------


class TestReportCatalog:
    """Tests verifying compliance reports appear in the catalog."""

    @pytest.mark.asyncio
    async def test_catalog_includes_compliance_types(self) -> None:
        """The report catalog should include soc2, iso27001, nist-csf."""
        from app.routers.reports import report_catalog

        # Mock the dependencies
        mock_user = MagicMock()
        mock_db = AsyncMock()

        catalog = await report_catalog(current_user=mock_user, db=mock_db)

        report_types = {item["type"] for item in catalog}
        assert "soc2" in report_types
        assert "iso27001" in report_types
        assert "nist-csf" in report_types

    @pytest.mark.asyncio
    async def test_catalog_compliance_tags(self) -> None:
        """Compliance reports should have 'compliance' tag."""
        from app.routers.reports import report_catalog

        mock_user = MagicMock()
        mock_db = AsyncMock()

        catalog = await report_catalog(current_user=mock_user, db=mock_db)

        compliance_reports = [r for r in catalog if r["type"] in ("soc2", "iso27001", "nist-csf")]
        for report in compliance_reports:
            assert "compliance" in report["tags"]


# ---------------------------------------------------------------------------
# Shared compliance helpers
# ---------------------------------------------------------------------------


class TestComplianceHelpers:
    """Tests for shared compliance report helper functions."""

    @pytest.mark.asyncio
    async def test_count_events(self) -> None:
        from app.services.compliance_report_service import _count_events

        cnt_row = _make_row(cnt=42)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(fetchone=MagicMock(return_value=cnt_row))
        )

        now = datetime.now(UTC)
        result = await _count_events(
            session, start=now - timedelta(days=30), end=now, action_filter="org.%"
        )
        assert result == 42

    @pytest.mark.asyncio
    async def test_count_events_with_org(self) -> None:
        from app.services.compliance_report_service import _count_events

        cnt_row = _make_row(cnt=10)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(fetchone=MagicMock(return_value=cnt_row))
        )

        now = datetime.now(UTC)
        result = await _count_events(
            session,
            start=now - timedelta(days=30),
            end=now,
            action_filter="org.%",
            org="my-org",
        )
        assert result == 10
        # Verify org was in params
        call_params = session.execute.call_args[0][1]
        assert call_params["org"] == "my-org"

    @pytest.mark.asyncio
    async def test_count_events_in_empty_actions(self) -> None:
        from app.services.compliance_report_service import _count_events_in

        session = AsyncMock()
        now = datetime.now(UTC)
        result = await _count_events_in(
            session, start=now - timedelta(days=30), end=now, actions=[]
        )
        assert result == 0
        session.execute.assert_not_called()

    @pytest.mark.asyncio
    async def test_distinct_actors(self) -> None:
        from app.services.compliance_report_service import _distinct_actors

        cnt_row = _make_row(cnt=15)
        session = AsyncMock()
        session.execute = AsyncMock(
            return_value=MagicMock(fetchone=MagicMock(return_value=cnt_row))
        )

        now = datetime.now(UTC)
        result = await _distinct_actors(
            session, start=now - timedelta(days=30), end=now, action_filter="%"
        )
        assert result == 15

    @pytest.mark.asyncio
    async def test_top_actions(self) -> None:
        from app.services.compliance_report_service import _top_actions

        rows = [
            _make_row(action="org.add_member", event_count=50),
            _make_row(action="org.remove_member", event_count=10),
        ]
        session = AsyncMock()
        session.execute = AsyncMock(return_value=MagicMock(fetchall=MagicMock(return_value=rows)))

        now = datetime.now(UTC)
        result = await _top_actions(
            session, start=now - timedelta(days=30), end=now, action_filter="org.%"
        )
        assert len(result) == 2
        assert result[0]["action"] == "org.add_member"
        assert result[0]["event_count"] == 50

    @pytest.mark.asyncio
    async def test_count_events_no_rows(self) -> None:
        from app.services.compliance_report_service import _count_events

        session = AsyncMock()
        session.execute = AsyncMock(return_value=MagicMock(fetchone=MagicMock(return_value=None)))

        now = datetime.now(UTC)
        result = await _count_events(
            session, start=now - timedelta(days=30), end=now, action_filter="org.%"
        )
        assert result == 0
