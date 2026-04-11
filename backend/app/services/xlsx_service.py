"""XLSX export service using openpyxl.

Generates Excel workbooks from report data with proper column types,
auto-width columns, and styled headers.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = structlog.get_logger(__name__)

# Header styling
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="24292E", end_color="24292E", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def generate_xlsx(
    data: list[dict[str, Any]],
    sheet_name: str = "Report",
    column_config: list[dict[str, str]] | None = None,
) -> bytes:
    """Generate an XLSX workbook from a list of dicts.

    Args:
        data: List of row dicts.  Keys become column headers if
            ``column_config`` is not provided.
        sheet_name: Name for the worksheet tab.
        column_config: Optional list of ``{"key": "field_name", "header": "Display Name"}``
            dicts to control column order and headers.  When omitted, all keys
            from the first row are used in order.

    Returns:
        The workbook as bytes suitable for ``StreamingResponse``.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel limits sheet names to 31 chars

    if not data:
        # Write a single "No data" row so the file is not completely empty
        ws.append(["No data available"])
        return _workbook_to_bytes(wb)

    # Determine columns
    if column_config:
        keys = [c["key"] for c in column_config]
        headers = [c.get("header", c["key"]) for c in column_config]
    else:
        keys = list(data[0].keys())
        headers = [_format_header(k) for k in keys]

    # Write header row
    ws.append(headers)

    # Write data rows, converting types as needed
    for row_dict in data:
        row_values: list[Any] = []
        for key in keys:
            value = row_dict.get(key)
            row_values.append(_coerce_value(value))
        ws.append(row_values)

    # Re-open workbook in normal mode to apply header styling and auto-width.
    # write_only mode streams efficiently but does not support cell styling.
    wb2 = Workbook()
    ws2 = wb2.active
    if ws2 is None:
        ws2 = wb2.create_sheet(title=sheet_name[:31])
    else:
        ws2.title = sheet_name[:31]

    # Re-write headers with style
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    # Re-write data
    for row_idx, row_dict in enumerate(data, 2):
        for col_idx, key in enumerate(keys, 1):
            value = row_dict.get(key)
            ws2.cell(row=row_idx, column=col_idx, value=_coerce_value(value))

    # Auto-width columns
    for col_idx in range(1, len(keys) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(len(data) + 2, 102)):  # Sample up to 100 rows
            cell = ws2.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted_width = min(max_len + 4, 60)
        ws2.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws2.freeze_panes = "A2"

    return _workbook_to_bytes(wb2)


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize a workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _format_header(key: str) -> str:
    """Convert snake_case key to Title Case header."""
    return key.replace("_", " ").title()


def _coerce_value(value: Any) -> Any:
    """Coerce Python values to Excel-friendly types."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value
    if isinstance(value, dict):
        # Flatten dicts to JSON-like string
        import json

        return json.dumps(value, default=str)
    if isinstance(value, list):
        import json

        return json.dumps(value, default=str)
    return value
